import csv
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from copy import copy

import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import InvalidFileException

TEMPLATE_SHEET = "Format"

CELL_MAP = {
    "office": "B3",
    "date": "B4",
    "user": "G4",
    "time": "B5",
    "method": "G5",
    "program": "A9",
    "dayreport": "A11",  # A11: フォントサイズを8固定
    "temp": "B13",
    "slack": "A16",      # A16: フォントサイズを8固定
}

ATTEND_VALUE = "出席"
ABSENT_SKIP_VALUE = "欠席時対応"

MSG_NOT_USERCASEDAILY = "userCaseDailyではありません。"
MSG_NOT_CASEDAILY = "caseDailyではありません。"
MSG_NOT_CSV = "csvファイルではありません。"
MSG_MONTH_MISMATCH = "userCaseDailyとcaseDailyの日時が合いません。"
MSG_CASE_NOT_SELECTED = "caseDailyが未選択です。"
MSG_OUTDIR_NOT_SELECTED = "出力先が未選択です。"
MSG_FILE_IN_USE = "ファイルにアクセスできません。別のプロセスが使用中です。"
MSG_TEMPLATE_NOT_FOUND = "java.io.FileNotFoundException.Sample_Format.xlsx(指定されたファイルが見つかりません。)"


def get_base_folder() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def is_csv(path: Path) -> bool:
    return path.suffix.lower() == ".csv"


def looks_like_userCaseDaily(path: Path) -> bool:
    return "userCaseDaily" in path.name


def looks_like_caseDaily(path: Path) -> bool:
    name = path.name
    return ("caseDaily" in name) or ("caseMonth" in name)


def extract_yyyymm_from_filename(path: Path) -> Optional[str]:
    m = re.search(r"_(\d{6})", path.name)
    return m.group(1) if m else None


def detect_encoding(path: Path) -> str:
    for enc in ("cp932", "shift_jis", "utf-8-sig", "utf-8"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                f.read(4096)
            return enc
        except Exception:
            pass
    raise RuntimeError(f"文字コード判定失敗: {path}")


def read_csv_dicts(path: Path) -> List[Dict[str, str]]:
    enc = detect_encoding(path)
    with path.open("r", encoding=enc, newline="") as f:
        reader = csv.DictReader(f)
        rows: List[Dict[str, str]] = []
        for r in reader:
            rows.append({(k or "").strip(): (v or "").strip() for k, v in r.items()})
        return rows


def normalize_date(s: str) -> str:
    return (s or "").strip().replace("-", "/")


# ===== 対応時間：「○時○分～○時○分」に寄せる =====
def parse_time_flexible(s: str) -> Optional[Tuple[int, int]]:
    s = (s or "").strip()
    if not s:
        return None

    patterns = [
        r"(\d{1,2}):(\d{2})(?::\d{2})?",
        r"(\d{1,2})時(\d{1,2})分",
    ]
    for pat in patterns:
        m = re.search(pat, s)
        if m:
            h = int(m.group(1))
            mi = int(m.group(2))
            if 0 <= h <= 23 and 0 <= mi <= 59:
                return (h, mi)
    return None


def format_time_range_jp(start: str, end: str) -> str:
    ps = parse_time_flexible(start)
    pe = parse_time_flexible(end)

    if ps and pe:
        sh, sm = ps
        eh, em = pe
        return f"{sh}時{sm:02d}分～{eh}時{em:02d}分"

    def fmt_one(p: Optional[Tuple[int, int]], raw: str) -> str:
        if p:
            h, m = p
            return f"{h}時{m:02d}分"
        return (raw or "").strip()

    left = fmt_one(ps, start)
    right = fmt_one(pe, end)

    if not left and not right:
        return ""
    if left and right:
        return f"{left}～{right}"
    return left or right
# ===============================================


def safe_sheet_name(name: str) -> str:
    for c in [":", "/", "\\", "?", "*", "[", "]"]:
        name = name.replace(c, "_")
    return name.strip()[:31]


def pick_date_column(daily_rows: List[Dict[str, str]]) -> str:
    candidates = ["日付", "年月日", "支援実施日"]
    keys = list(daily_rows[0].keys())
    for c in candidates:
        if c in keys:
            return c
    return keys[0]


def pick_daily_note(daily: Dict[str, str]) -> str:
    candidates = ["備考", "備考欄", "本人との連絡", "連絡", "連絡事項"]
    for c in candidates:
        v = (daily.get(c) or "").strip()
        if v:
            return v
    return ""


def build_program(d: Dict[str, str]) -> str:
    out: List[str] = []

    def add(p, detail):
        p = (p or "").strip()
        detail = (detail or "").strip()
        if p or detail:
            out.append(p + ("\n" + detail if (p and detail) else detail))

    add(d.get("午前のプログラム", ""), d.get("午前のプログラム詳細", ""))
    add(d.get("午後1のプログラム", ""), d.get("午後1のプログラム詳細", ""))
    add(d.get("午後2のプログラム", ""), d.get("午後2のプログラム詳細", ""))
    add(d.get("終日のプログラム", ""), d.get("終日のプログラム詳細", ""))
    return "\n".join(out)


def normalize_method(raw: str) -> str:
    raw = raw or ""
    if "在宅" in raw and "通所" not in raw:
        return "利用者宅"
    return "事業所"


def format_contact_text(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""

    parts = re.split(r"(\b\d{1,2}:\d{2}\b)", text)
    if len(parts) == 1:
        body = text
        return (body[:30] + "・・・・") if len(body) > 30 else body

    lines: List[str] = []
    i = 0
    while i < len(parts):
        seg = parts[i]
        if re.fullmatch(r"\b\d{1,2}:\d{2}\b", seg or ""):
            t = seg
            msg = (parts[i + 1] if i + 1 < len(parts) else "").strip()
            if len(msg) > 30:
                msg = msg[:30] + "・・・・"
            lines.append(f"{t} {msg}".rstrip())
            i += 2
        else:
            i += 1

    return "\n".join([ln for ln in lines if ln])


def remove_sample_sheets(wb) -> None:
    targets = [name for name in wb.sheetnames if "sample" in name.lower()]
    for name in targets:
        del wb[name]


def ask_paths() -> Tuple[Optional[Path], Optional[Path], Optional[Path]]:
    root = tk.Tk()
    root.withdraw()

    user_path_str = filedialog.askopenfilename(title="userCaseDailyを選択", filetypes=[("CSV", "*.*")])
    if not user_path_str:
        return None, None, None
    user_path = Path(user_path_str)

    if not is_csv(user_path):
        messagebox.showerror("エラー", MSG_NOT_CSV)
        return None, None, None
    if not looks_like_userCaseDaily(user_path):
        messagebox.showerror("エラー", MSG_NOT_USERCASEDAILY)
        return None, None, None

    case_path_str = filedialog.askopenfilename(title="caseDailyを選択", filetypes=[("CSV", "*.*")])
    if not case_path_str:
        messagebox.showerror("エラー", MSG_CASE_NOT_SELECTED)
        return None, None, None
    case_path = Path(case_path_str)

    if not is_csv(case_path):
        messagebox.showerror("エラー", MSG_NOT_CSV)
        return None, None, None
    if not looks_like_caseDaily(case_path):
        messagebox.showerror("エラー", MSG_NOT_CASEDAILY)
        return None, None, None

    outdir_str = filedialog.askdirectory(title="出力先フォルダを選択")
    if not outdir_str:
        messagebox.showerror("エラー", MSG_OUTDIR_NOT_SELECTED)
        return None, None, None
    outdir = Path(outdir_str)

    return user_path, case_path, outdir


def ensure_same_month(user_path: Path, case_path: Path) -> None:
    u = extract_yyyymm_from_filename(user_path)
    c = extract_yyyymm_from_filename(case_path)
    if u and c and (u != c):
        raise ValueError(MSG_MONTH_MISMATCH)


def build_output_filename(case_rows: List[Dict[str, str]], yyyymm: Optional[str]) -> str:
    name = (case_rows[0].get("氏名") or "").strip() or "名前未設定"
    if not yyyymm:
        d = normalize_date(case_rows[0].get("年月日", ""))
        m = re.match(r"^(\d{4})/(\d{1,2})", d)
        yyyymm = f"{m.group(1)}{int(m.group(2)):02d}" if m else "YYYYMM"
    return f"{name}_{yyyymm}_サービス支援記録.xlsx"


def load_template_or_fail(base: Path) -> Path:
    tpl = base / "Sample_Format.xlsx"
    if not tpl.exists():
        raise FileNotFoundError(MSG_TEMPLATE_NOT_FOUND)
    return tpl


def set_font_size_only(ws, addr: str, size: int):
    """
    テンプレの形は変えず、フォントサイズだけ変更する。
    （太字・色・フォント名などは保持）
    """
    c = ws[addr]
    f = c.font
    nf = copy(f)
    nf.size = size
    c.font = nf


def generate(user_csv: Path, case_csv: Path, outdir: Path) -> Path:
    base = get_base_folder()
    template_path = load_template_or_fail(base)

    ensure_same_month(user_csv, case_csv)

    case_rows = read_csv_dicts(case_csv)
    daily_rows = read_csv_dicts(user_csv)
    if not case_rows:
        raise RuntimeError("caseDailyが空です。")
    if not daily_rows:
        raise RuntimeError("userCaseDailyが空です。")

    yyyymm = extract_yyyymm_from_filename(case_csv) or extract_yyyymm_from_filename(user_csv)
    out_name = build_output_filename(case_rows, yyyymm)
    out_path = outdir / out_name

    if out_path.exists():
        msg = f"このフォルダーには’{out_name}’は存在します。上書きしますか？"
        if not messagebox.askyesno("確認", msg):
            raise RuntimeError("キャンセルしました。")

    try:
        wb = load_workbook(template_path)
    except (InvalidFileException, Exception) as e:
        raise RuntimeError(f"テンプレ読み込み失敗: {e}")

    remove_sample_sheets(wb)

    if TEMPLATE_SHEET not in wb.sheetnames:
        raise RuntimeError(f"テンプレに '{TEMPLATE_SHEET}' シートがありません。")
    tpl = wb[TEMPLATE_SHEET]

    date_col = pick_date_column(daily_rows)
    daily_by_date: Dict[str, Dict[str, str]] = {}
    for r in daily_rows:
        daily_by_date[normalize_date(r.get(date_col, ""))] = r

    required = ["事業所名", "氏名", "年月日", "出欠等", "実績開始時間", "実績終了時間"]
    for c in required:
        if c not in case_rows[0]:
            raise RuntimeError(f"caseDailyに必須列がありません: {c}")

    for r in case_rows:
        status = (r.get("出欠等", "") or "").strip()
        if status == ABSENT_SKIP_VALUE:
            continue
        if status != ATTEND_VALUE:
            continue

        date = normalize_date(r.get("年月日", ""))
        if not date:
            continue

        daily = daily_by_date.get(date, {})

        sheet_base = f"{date.replace('/','')[:8]}_{(r.get('氏名','') or '').strip()}"
        sheet_name = safe_sheet_name(sheet_base)
        if sheet_name in wb.sheetnames:
            k = 2
            while True:
                cand = safe_sheet_name(f"{sheet_base}_{k}")
                if cand not in wb.sheetnames:
                    sheet_name = cand
                    break
                k += 1

        ws = wb.copy_worksheet(tpl)
        ws.title = sheet_name

        ws[CELL_MAP["office"]].value = r.get("事業所名", "")
        ws[CELL_MAP["date"]].value = date
        ws[CELL_MAP["user"]].value = r.get("氏名", "")

        ws[CELL_MAP["time"]].value = format_time_range_jp(
            r.get("実績開始時間", ""),
            r.get("実績終了時間", "")
        )

        method_cell = ws[CELL_MAP["method"]]
        method_cell.value = normalize_method(r.get("実績記録票備考欄", ""))
        method_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws[CELL_MAP["program"]].value = build_program(daily)

        # A11：値を入れて、フォントサイズだけ 12→8
        ws[CELL_MAP["dayreport"]].value = r.get("日報", "")
        set_font_size_only(ws, CELL_MAP["dayreport"], 8)

        temp = (daily.get("体温", "") or "").strip()
        ws[CELL_MAP["temp"]].value = "未検温" if temp == "" else f"{temp}℃"

        daily_note = pick_daily_note(daily)
        cm_note = (r.get("備考") or r.get("実績記録票備考欄") or "").strip()
        raw_contact = daily_note or cm_note

        # A16：値を入れて、フォントサイズだけ 12→8
        ws[CELL_MAP["slack"]].value = format_contact_text(raw_contact)
        set_font_size_only(ws, CELL_MAP["slack"], 8)

    remove_sample_sheets(wb)

    try:
        wb.save(out_path)
    except PermissionError:
        raise PermissionError(MSG_FILE_IN_USE)

    return out_path


def main():
    root = tk.Tk()
    root.withdraw()

    user_path, case_path, outdir = ask_paths()
    if user_path is None and case_path is None and outdir is None:
        return

    try:
        out_path = generate(user_path, case_path, outdir)
        messagebox.showinfo("完了", f"保存しました。\n{out_path}")
    except Exception as e:
        messagebox.showerror("エラー", str(e))


if __name__ == "__main__":
    main()
